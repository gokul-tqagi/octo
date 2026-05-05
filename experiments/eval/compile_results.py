# compile_results.py: Build Excel spreadsheet of all Octo finetuning experiments.
# compile_results.py: Auto-discovers eval results and configs, appends new runs incrementally.

"""
Compile eval metrics from all training runs into an Excel spreadsheet.

Scans eval_output dirs for eval_metrics.json, matches with finetune configs,
and builds a comparison table with key metrics + design choices.

Usage:
    # Rebuild full spreadsheet from all results
    python3 experiments/eval/compile_results.py --data_dir /home/gokul/data --config_dir experiments/configs --output results.xlsx

    # Add a single new run (appends if sheet exists)
    python3 experiments/eval/compile_results.py --data_dir /home/gokul/data --config_dir experiments/configs --output results.xlsx --run lipbalm_10hz
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import numpy as np

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl required: pip install openpyxl")
    sys.exit(1)

try:
    import yaml
except ImportError:
    print("pyyaml required: pip install pyyaml")
    sys.exit(1)


# Map from eval_output dir name -> finetune config filename
# Handles cases where the checkpoint dir name doesn't exactly match the config
CONFIG_MAP = {
    "marker_pick_9999": "finetune_marker_pick.yaml",
    "marker_pick_19999": "finetune_marker_pick.yaml",
    "marker_pick_40k": "finetune_marker_pick_40k.yaml",
    "marker_pick_full": "finetune_marker_pick_full.yaml",
    "marker_pick_10hz": "finetune_marker_pick_10hz.yaml",
    "lipbalm_20k": "finetune_lipbalm.yaml",
    "lipbalm_10hz": "finetune_lipbalm_10hz.yaml",
}

# Human-readable run descriptions
RUN_COMMENTS = {
    "marker_pick_9999": "Early checkpoint (10k steps). Undertrained baseline.",
    "marker_pick_19999": "Head-only 20k steps at 5Hz. First converged marker_pick run.",
    "marker_pick_40k": "Continued head-only to 40k steps. Marginal gain over 20k.",
    "marker_pick_full": "Full finetune from 40k ckpt. No improvement — slight overfit.",
    "marker_pick_10hz": "Head-only 20k at 10Hz. 2x data, half step size. Best marker_pick.",
    "lipbalm_20k": "Head-only 20k steps at 5Hz. Baseline lipbalm run.",
    "lipbalm_10hz": "Head-only 20k at 10Hz. 2x data, half step size. Best lipbalm.",
}


def parse_config(config_path):
    """Extract training hyperparams from a finetune YAML config."""
    if not os.path.exists(config_path):
        return {}
    with open(config_path) as f:
        cfg = yaml.safe_load(f)
    return {
        "dataset_name": cfg.get("dataset_name", ""),
        "num_steps": cfg.get("num_steps", ""),
        "learning_rate": cfg.get("learning_rate", ""),
        "batch_size": cfg.get("batch_size", ""),
        "freeze_transformer": cfg.get("freeze_transformer", ""),
        "pretrained_path": cfg.get("pretrained_path", ""),
    }


def infer_config(run_name, config_dir):
    """Find the matching finetune config for a run."""
    # Check explicit map first
    if run_name in CONFIG_MAP:
        path = os.path.join(config_dir, CONFIG_MAP[run_name])
        if os.path.exists(path):
            return path

    # Try direct name match
    for pattern in [f"finetune_{run_name}.yaml", f"finetune_{run_name.replace('_10hz', '_10hz')}.yaml"]:
        path = os.path.join(config_dir, pattern)
        if os.path.exists(path):
            return path

    return None


def infer_data_hz(run_name, config):
    """Infer sampling rate from run name or dataset name."""
    if "10hz" in run_name.lower():
        return 10.0
    ds = config.get("dataset_name", "")
    if "10hz" in ds:
        return 10.0
    return 5.0


def infer_task(run_name):
    """Infer task name from run name."""
    if "marker" in run_name:
        return "marker_pick"
    if "lipbalm" in run_name:
        return "lipbalm_pick"
    if "aloha" in run_name:
        return "pack_red_bag"
    return run_name


def infer_total_steps(run_name, config):
    """Infer cumulative training steps (handles continued training)."""
    steps = config.get("num_steps", 20000)
    pretrained = config.get("pretrained_path", "")
    if "40k" in run_name:
        return 40000
    if "full" in run_name:
        return 60000
    if "9999" in run_name:
        return 10000
    if "19999" in run_name:
        return 20000
    if "/checkpoints/" in str(pretrained):
        # Continued from a checkpoint — this is additive
        return steps  # Just report this run's steps
    return steps


def load_eval_metrics(eval_dir):
    """Load eval_metrics.json and compute aggregates."""
    metrics_path = os.path.join(eval_dir, "eval_metrics.json")
    if not os.path.exists(metrics_path):
        return None

    with open(metrics_path) as f:
        episodes = json.load(f)

    if not episodes:
        return None

    keys = [
        "pos_error_mean_mm", "pos_error_max_mm",
        "rot_error_mean_deg", "rot_error_max_deg",
        "traj_error_mean_mm", "traj_error_final_mm",
    ]

    agg = {"n_episodes": len(episodes)}
    for k in keys:
        vals = [ep[k] for ep in episodes if k in ep]
        if vals:
            agg[f"{k}_mean"] = float(np.mean(vals))
            agg[f"{k}_std"] = float(np.std(vals))

    return agg


def find_training_log(data_dir, run_name):
    """Find the training log for a run."""
    # Check common locations
    for subdir in ["checkpoints", "eval_output"]:
        base = os.path.join(data_dir, subdir, run_name)
        if os.path.isdir(base):
            for f in os.listdir(base):
                if f.startswith("train_") and f.endswith(".log"):
                    return os.path.join(base, f)
    return None


def parse_final_loss(log_path):
    """Read last loss value from training log."""
    if not log_path or not os.path.exists(log_path):
        return None
    with open(log_path) as f:
        lines = f.readlines()
    for line in reversed(lines):
        line = line.strip()
        if "," in line:
            parts = line.split(",")
            if len(parts) == 2:
                try:
                    return float(parts[1])
                except ValueError:
                    continue
    return None


def discover_runs(data_dir):
    """Find all eval output directories."""
    eval_dir = os.path.join(data_dir, "eval_output")
    if not os.path.isdir(eval_dir):
        return []
    runs = []
    for name in sorted(os.listdir(eval_dir)):
        metrics_path = os.path.join(eval_dir, name, "eval_metrics.json")
        if os.path.exists(metrics_path):
            runs.append(name)
    return runs


def build_row(run_name, data_dir, config_dir):
    """Build one spreadsheet row for a run."""
    eval_dir = os.path.join(data_dir, "eval_output", run_name)
    metrics = load_eval_metrics(eval_dir)
    if metrics is None:
        return None

    config_path = infer_config(run_name, config_dir)
    config = parse_config(config_path) if config_path else {}

    log_path = find_training_log(data_dir, run_name)
    final_loss = parse_final_loss(log_path)

    data_hz = infer_data_hz(run_name, config)
    task = infer_task(run_name)
    total_steps = infer_total_steps(run_name, config)
    finetune_mode = "full" if config.get("freeze_transformer") is False else "head_only"
    comment = RUN_COMMENTS.get(run_name, "")

    return {
        "run_name": run_name,
        "task": task,
        "data_hz": data_hz,
        "finetune_mode": finetune_mode,
        "total_steps": total_steps,
        "learning_rate": config.get("learning_rate", ""),
        "batch_size": config.get("batch_size", ""),
        "final_loss": final_loss,
        "n_eval_episodes": metrics["n_episodes"],
        "pos_err_mean_mm": metrics.get("pos_error_mean_mm_mean"),
        "pos_err_max_mm": metrics.get("pos_error_max_mm_mean"),
        "rot_err_mean_deg": metrics.get("rot_error_mean_deg_mean"),
        "rot_err_max_deg": metrics.get("rot_error_max_deg_mean"),
        "traj_err_mean_mm": metrics.get("traj_error_mean_mm_mean"),
        "traj_err_final_mm": metrics.get("traj_error_final_mm_mean"),
        "pos_err_std": metrics.get("pos_error_mean_mm_std"),
        "traj_final_std": metrics.get("traj_error_final_mm_std"),
        "eval_artifacts": os.path.abspath(eval_dir),
        "config_file": os.path.basename(config_path) if config_path else "",
        "comments": comment,
    }


# Column definitions: (header, key, width, fmt)
COLUMNS = [
    ("Run Name", "run_name", 22, None),
    ("Task", "task", 14, None),
    ("Hz", "data_hz", 6, "0.0"),
    ("Mode", "finetune_mode", 11, None),
    ("Steps", "total_steps", 8, "#,##0"),
    ("LR", "learning_rate", 10, "0.0E+0"),
    ("Batch", "batch_size", 7, "0"),
    ("Final Loss", "final_loss", 11, "0.000"),
    ("Eval Eps", "n_eval_episodes", 9, "0"),
    ("Pos Mean (mm)", "pos_err_mean_mm", 15, "0.0"),
    ("Pos Max (mm)", "pos_err_max_mm", 14, "0.0"),
    ("Pos Std (mm)", "pos_err_std", 13, "0.0"),
    ("Rot Mean (deg)", "rot_err_mean_deg", 15, "0.0"),
    ("Rot Max (deg)", "rot_err_max_deg", 14, "0.0"),
    ("Traj Mean (mm)", "traj_err_mean_mm", 15, "0.0"),
    ("Traj Final (mm)", "traj_err_final_mm", 16, "0.0"),
    ("Traj Final Std", "traj_final_std", 14, "0.0"),
    ("Artifacts Path", "eval_artifacts", 50, None),
    ("Config", "config_file", 30, None),
    ("Comments", "comments", 55, None),
]


def write_excel(rows, output_path):
    """Write or overwrite the Excel spreadsheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Octo Finetune Results"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, (header, _, width, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Group rows by task
    task_order = ["marker_pick", "lipbalm_pick", "pack_red_bag"]
    sorted_rows = sorted(rows, key=lambda r: (
        task_order.index(r["task"]) if r["task"] in task_order else 99,
        r["data_hz"],
        r["total_steps"],
    ))

    # Alternating row colors per task
    task_colors = {
        "marker_pick": ("D6E4F0", "EDF2F9"),
        "lipbalm_pick": ("E2EFDA", "F0F5EA"),
        "pack_red_bag": ("FCE4D6", "FDF0E8"),
    }

    # Best values per task for conditional formatting
    best_per_task = {}
    for row in sorted_rows:
        task = row["task"]
        if task not in best_per_task:
            best_per_task[task] = {}
        for key in ["pos_err_mean_mm", "rot_err_mean_deg", "traj_err_final_mm"]:
            val = row.get(key)
            if val is not None:
                if key not in best_per_task[task] or val < best_per_task[task][key]:
                    best_per_task[task][key] = val

    best_font = Font(bold=True, color="006100")
    best_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    current_task = None
    row_within_task = 0

    for row_idx, row_data in enumerate(sorted_rows, 2):
        task = row_data["task"]
        if task != current_task:
            current_task = task
            row_within_task = 0
        row_within_task += 1

        colors = task_colors.get(task, ("FFFFFF", "F5F5F5"))
        bg_color = colors[0] if row_within_task % 2 == 1 else colors[1]
        row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

        for col_idx, (_, key, _, fmt) in enumerate(COLUMNS, 1):
            val = row_data.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center")

            if fmt and val not in (None, ""):
                cell.number_format = fmt

            # Highlight best values
            if key in ["pos_err_mean_mm", "rot_err_mean_deg", "traj_err_final_mm"]:
                if val is not None and task in best_per_task:
                    if abs(val - best_per_task[task].get(key, float("inf"))) < 0.01:
                        cell.font = best_font
                        cell.fill = best_fill

    # Freeze top row
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Add summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Octo Finetuning Experiment Tracker").font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws2.cell(row=3, column=1, value=f"Total runs: {len(rows)}")
    ws2.cell(row=5, column=1, value="Key Findings:").font = Font(bold=True, size=11)
    findings = [
        "10Hz sampling halves per-step errors vs 5Hz across both tasks",
        "Head-only finetuning sufficient for 50-75 episode datasets",
        "Full finetune from head-only checkpoint shows no gain (possible overfit)",
        "Diminishing returns beyond 20k steps for head-only",
        "Lipbalm task has lower trajectory error than marker_pick (shorter episodes)",
    ]
    for i, finding in enumerate(findings):
        ws2.cell(row=6 + i, column=1, value=f"  {i+1}. {finding}")

    wb.save(output_path)
    print(f"Saved: {output_path} ({len(rows)} runs)")


def main():
    parser = argparse.ArgumentParser(description="Compile Octo finetune results to Excel")
    parser.add_argument("--data_dir", required=True, help="Base data dir with eval_output/ and checkpoints/")
    parser.add_argument("--config_dir", required=True, help="Dir with finetune_*.yaml configs")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    parser.add_argument("--run", help="Only add/update this specific run (appends if sheet exists)")
    args = parser.parse_args()

    config_dir = os.path.abspath(args.config_dir)

    if args.run:
        runs = [args.run]
    else:
        runs = discover_runs(args.data_dir)

    if not runs:
        print("No eval results found.")
        sys.exit(1)

    print(f"Found {len(runs)} runs: {', '.join(runs)}")

    # If appending a single run to existing sheet, load existing rows first
    existing_rows = []
    if args.run and os.path.exists(args.output):
        wb = openpyxl.load_workbook(args.output)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        key_to_col = {}
        for col_def in COLUMNS:
            header_name = col_def[0]
            key_name = col_def[1]
            if header_name in headers:
                key_to_col[key_name] = headers.index(header_name)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            row_dict = {}
            for key_name, col_idx in key_to_col.items():
                row_dict[key_name] = row[col_idx]
            # Skip the run we're updating
            if row_dict.get("run_name") != args.run:
                existing_rows.append(row_dict)
        wb.close()

    # Build rows for requested runs
    new_rows = []
    for run_name in runs:
        row = build_row(run_name, args.data_dir, config_dir)
        if row:
            new_rows.append(row)
            print(f"  {run_name}: pos={row['pos_err_mean_mm']:.1f}mm, traj_final={row['traj_err_final_mm']:.1f}mm")
        else:
            print(f"  {run_name}: no metrics found, skipping")

    all_rows = existing_rows + new_rows
    write_excel(all_rows, args.output)


if __name__ == "__main__":
    main()
